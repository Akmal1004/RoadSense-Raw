"""Excel Report Generator for Selenium E2E Automation Framework."""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from automation.config.config import Config

class ExcelReporter:
    """Generates enterprise Excel workbooks for E2E execution results."""

    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    PASS_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    PASS_FONT = Font(name="Segoe UI", size=10, color="276A3C", bold=True)

    FAIL_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    FAIL_FONT = Font(name="Segoe UI", size=10, color="C00000", bold=True)

    SKIP_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    SKIP_FONT = Font(name="Segoe UI", size=10, color="B25900", bold=True)

    BORDER_THIN = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    @classmethod
    def generate_all_excel_reports(cls, test_results: list, metrics: dict):
        """Generate all 4 required Excel report workbooks."""
        os.makedirs(Config.EXCEL_DIR, exist_ok=True)

        cls._generate_master_report(test_results, metrics)
        cls._generate_passed_report(test_results)
        cls._generate_failed_report(test_results)
        cls._generate_summary_report(metrics)

    @classmethod
    def _generate_master_report(cls, test_results: list, metrics: dict):
        wb = openpyxl.Workbook()

        # Sheet 1: Executed Test Cases
        ws_all = wb.active
        ws_all.title = "Executed Test Cases"
        headers_all = ["Test ID", "Module", "Test Name", "Status", "Execution Time (s)", "Priority", "Failure Reason"]
        cls._build_table_sheet(ws_all, headers_all, test_results)

        # Sheet 2: Passed Tests
        ws_pass = wb.create_sheet("Passed Tests")
        passed_results = [r for r in test_results if r["status"] == "PASS"]
        cls._build_table_sheet(ws_pass, headers_all, passed_results)

        # Sheet 3: Failed Tests
        ws_fail = wb.create_sheet("Failed Tests")
        failed_results = [r for r in test_results if r["status"] == "FAIL"]
        cls._build_table_sheet(ws_fail, headers_all, failed_results)

        # Sheet 4: Skipped Tests
        ws_skip = wb.create_sheet("Skipped Tests")
        skipped_results = [r for r in test_results if r["status"] in ["SKIPPED", "BLOCKED"]]
        cls._build_table_sheet(ws_skip, headers_all, skipped_results)

        # Sheet 5: Execution Metrics
        ws_metrics = wb.create_sheet("Execution Metrics")
        cls._build_metrics_sheet(ws_metrics, metrics)

        # Sheet 6: Defect Summary
        ws_defects = wb.create_sheet("Defect Summary")
        cls._build_defects_sheet(ws_defects, failed_results)

        file_path = os.path.join(Config.EXCEL_DIR, "Automation_Test_Report.xlsx")
        wb.save(file_path)
        print(f"[SUCCESS] Generated Excel Master Report: {file_path}")

    @classmethod
    def _generate_passed_report(cls, test_results: list):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Passed Test Cases"
        passed_results = [r for r in test_results if r["status"] == "PASS"]
        headers = ["Test ID", "Module", "Test Name", "Status", "Execution Time (s)", "Priority"]
        cls._build_table_sheet(ws, headers, passed_results)
        wb.save(os.path.join(Config.EXCEL_DIR, "Passed_Test_Cases.xlsx"))

    @classmethod
    def _generate_failed_report(cls, test_results: list):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Failed Test Cases"
        failed_results = [r for r in test_results if r["status"] == "FAIL"]
        headers = ["Test ID", "Module", "Test Name", "Status", "Execution Time (s)", "Priority", "Failure Reason"]
        cls._build_table_sheet(ws, headers, failed_results)
        wb.save(os.path.join(Config.EXCEL_DIR, "Failed_Test_Cases.xlsx"))

    @classmethod
    def _generate_summary_report(cls, metrics: dict):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Executive Summary"
        cls._build_metrics_sheet(ws, metrics)
        wb.save(os.path.join(Config.EXCEL_DIR, "Summary_Report.xlsx"))

    @classmethod
    def _build_table_sheet(cls, ws, headers: list, data: list):
        ws.append(headers)
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = cls.HEADER_FILL
            cell.font = cls.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, item in enumerate(data, 2):
            row_data = [
                item.get("test_id", f"TC-{row_idx:03d}"),
                item.get("module", "General"),
                item.get("test_name", "Test"),
                item.get("status", "PASS"),
                item.get("duration", 0.0),
                item.get("priority", "P2"),
                item.get("error_message", "")
            ]
            ws.append(row_data[:len(headers)])

            status_cell = ws.cell(row=row_idx, column=4)
            if item.get("status") == "PASS":
                status_cell.fill = cls.PASS_FILL
                status_cell.font = cls.PASS_FONT
            elif item.get("status") == "FAIL":
                status_cell.fill = cls.FAIL_FILL
                status_cell.font = cls.FAIL_FONT
            else:
                status_cell.fill = cls.SKIP_FILL
                status_cell.font = cls.SKIP_FONT

            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = cls.BORDER_THIN

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)

    @classmethod
    def _build_metrics_sheet(cls, ws, metrics: dict):
        ws.append(["Metric Parameter", "Value"])
        ws.cell(row=1, column=1).fill = cls.HEADER_FILL
        ws.cell(row=1, column=1).font = cls.HEADER_FONT
        ws.cell(row=1, column=2).fill = cls.HEADER_FILL
        ws.cell(row=1, column=2).font = cls.HEADER_FONT

        metric_rows = [
            ("Target Environment URL", metrics.get("base_url", Config.BASE_URL)),
            ("Total Executed Test Cases", metrics.get("total", 0)),
            ("Passed Test Cases", metrics.get("passed", 0)),
            ("Failed Test Cases", metrics.get("failed", 0)),
            ("Skipped / Blocked Cases", metrics.get("skipped", 0)),
            ("Pass Rate (%)", f"{metrics.get('pass_rate', 0.0):.2f}%"),
            ("Total Execution Time (seconds)", f"{metrics.get('total_duration', 0.0):.2f}s"),
            ("Framework Engine", "Selenium WebDriver 4 + Pytest"),
            ("Browser Engine", f"Google Chrome (Headless)")
        ]
        for idx, (k, v) in enumerate(metric_rows, 2):
            ws.append([k, v])
            ws.cell(row=idx, column=1).border = cls.BORDER_THIN
            ws.cell(row=idx, column=2).border = cls.BORDER_THIN

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 50

    @classmethod
    def _build_defects_sheet(cls, ws, failed_results: list):
        headers = ["Defect ID", "Test Case ID", "Module", "Test Name", "Priority", "Failure Summary", "Screenshot Log"]
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = cls.HEADER_FILL
            cell.font = cls.HEADER_FONT

        for idx, item in enumerate(failed_results, 2):
            ws.append([
                f"DEF-{idx-1:03d}",
                item.get("test_id", "TC-000"),
                item.get("module", "General"),
                item.get("test_name", "Test"),
                item.get("priority", "P1"),
                item.get("error_message", "Assertion failed"),
                item.get("screenshot_path", "")
            ])
            for col in range(1, len(headers) + 1):
                ws.cell(row=idx, column=col).border = cls.BORDER_THIN

        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = 25
